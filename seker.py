#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error, r2_score
from sklearn import model_selection
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor
import openpyxl


# In[ ]:


from warnings import filterwarnings
filterwarnings("ignore")


# In[ ]:


df =  pd.read_excel('//MPSTLPSRV2//TalepTahminPaylasim2//model//verison.xlsx')
df1 =  pd.read_excel('//MPSTLPSRV2//TalepTahminPaylasim2/model//üretim.xlsx')


# In[ ]:


df.info()
df1.info()


# In[ ]:


df = df.dropna()


# In[ ]:


sehir=df["Şehir"]
y_üret=df["Radiation"]


# In[ ]:


toplam=0
toplaagg=0
aksaray_agırlık=0.1
konya_agırlık=0.5
nigde_agırlık=0.1
karaman_agırlık=0.1
nevsehir_agırlık=0.1
kırsehir_agırlık=0.1
toplaagırlık=aksaray_agırlık+nigde_agırlık+konya_agırlık+karaman_agırlık+nevsehir_agırlık+kırsehir_agırlık
print(aksaray_agırlık)
print(nigde_agırlık)
print(konya_agırlık)
print(karaman_agırlık)
print(nevsehir_agırlık)
print(kırsehir_agırlık)


# In[ ]:


print(toplaagırlık)


# In[ ]:


k=df[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]


# In[ ]:


for v in range(len(y_üret.index)): 
    if (sehir[v] == 'AKSARAY') :
        k[v:v+1]=k[v:v+1]*aksaray_agırlık       
    elif(sehir[v] == 'NİĞDE') :
        k[v:v+1]=k[v:v+1]*nigde_agırlık
    elif(sehir[v] == 'KONYA') :
        k[v:v+1]=k[v:v+1]*konya_agırlık
    elif(sehir[v] == 'KARAMAN') :
        k[v:v+1]=k[v:v+1]*karaman_agırlık
    elif(sehir[v] == 'KIRŞEHİR') :
        k[v:v+1]=k[v:v+1]*kırsehir_agırlık
    elif(sehir[v] == 'NEVŞEHİR') :
        k[v:v+1]=k[v:v+1]*nevsehir_agırlık


# In[ ]:


df[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]=k[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]
table = pd.pivot_table(df,index=['Tarih'],aggfunc=np.sum)


# In[ ]:


df1.index=table.index


# In[ ]:


table=table.join(df1)


# In[ ]:


y = table.loc[:,["GES NET"]]
c=table.loc[:,["Radiation","Effective Cloud","Medium Cloud","Solar Power","Precipitation","Low Cloud","Clear Radiation","Temperature"]]
print('X Shape: ', c.shape)
print('Y Shape: ', y.shape)


# In[ ]:


c_train, c_test, y_train, y_test = train_test_split(c,
                                                    y,
                                                    test_size = 0.00015,
                                                    random_state= 42)


# In[ ]:


rf_model= RandomForestRegressor(n_estimators = 10, random_state = 42)
rf_tuned = rf_model.fit(c,y.values.ravel())


# In[ ]:


rf_model = RandomForestRegressor(random_state= 42,
                                 max_depth=8, 
                                 max_features=7,
                                 min_samples_split= 10,
                                 n_estimators= 20000)
rf_tuned = rf_model.fit(c_train,y_train)


# In[ ]:


import pypyodbc
db = pypyodbc.connect(
   'Driver={SQL Server};'
    'Server=10.242.1.135;'
    'Database=Mepas;'
    'UID=TalepTahminGoruntuleme;'
    'PWD=Mepas@951;')
imlec = db.cursor()


# In[ ]:


df = pd.read_sql_query("SELECT Sehir,Tarih,tip, Sirket,Temperature,Precipitation,WindSpeed,LowCloud,MediumCloud,HighCloud,EffectiveCloud,Radiation,ClearRadiation,SolarPower From K3HavaTahmini",db)


# In[ ]:


hava=df[df['tip'] == 'Tahmin']


# In[ ]:


hava_tahmin=df[df['tip'] == 'Tahmin']


# In[ ]:


hava_tahmin.rename(columns ={'sehir':'Şehir','tarih':'Tarih','radiation':'Radiation','temperature':'Temperature','precipitation':'Precipitation','windspeed':'Wind Speed','lowcloud':'Low Cloud','mediumcloud':'Medium Cloud','highcloud':'High Cloud','effectivecloud':'Effective Cloud','clearradiation':'Clear Radiation','solarpower':'Solar Power'},inplace=True)


# In[ ]:


hava_tahmin.reset_index(drop=True,inplace=True)


# In[ ]:


sehir_tah=hava_tahmin["Şehir"]
ke=hava_tahmin[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]
for v in range(len(hava_tahmin.index)): 
    if (sehir_tah[v] == 'AKSARAY') :
        ke[v:v+1]=ke[v:v+1]*aksaray_agırlık
    elif(sehir_tah[v] == 'NİĞDE') :
        ke[v:v+1]=ke[v:v+1]*nigde_agırlık
    elif(sehir_tah[v]  == 'KONYA') :
        ke[v:v+1]=ke[v:v+1]*konya_agırlık
    elif(sehir_tah[v]  == 'KARAMAN') :
        ke[v:v+1]=ke[v:v+1]*karaman_agırlık
    elif(sehir_tah[v]  == 'KIRŞEHİR') :
        ke[v:v+1]=ke[v:v+1]*kırsehir_agırlık
    elif(sehir_tah[v]  == 'NEVŞEHİR') :
        ke[v:v+1]=ke[v:v+1]*nevsehir_agırlık


# In[ ]:


hava_tahmin[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]=ke[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]


# In[ ]:


table_tah = pd.pivot_table(hava_tahmin,index=['Tarih'],aggfunc=np.sum)


# In[ ]:


c_tahmin=table_tah.loc[:,["Radiation","Effective Cloud","Medium Cloud","Solar Power","Precipitation","Low Cloud","Clear Radiation","Temperature"]]
c_tahmin=c_tahmin.reset_index(drop=True)


# In[ ]:


import datetime
tarih =datetime.datetime.now()
zaman_damgası = datetime.datetime.timestamp(tarih)
zaman_damgası
z=str(zaman_damgası) + ".xlsx"


# In[ ]:


wb = openpyxl.Workbook()
sayfa = wb.active
for v in range(len(c_tahmin["Radiation"].index)):     
    y_pred =rf_tuned.predict(c_tahmin.loc[[v],["Radiation","Effective Cloud","Medium Cloud","Solar Power","Precipitation","Low Cloud","Clear Radiation","Temperature"]]) 
    r = v + 1
    sayfa.cell(row = r, column = 2).value = float(y_pred)
    sayfa.cell(row = r, column = 1).value = table_tah.index[v]
wb.save(z)
wb.close()


# In[ ]:


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
sender = 'info@mepasenerji.com'
receivers = ['enerjipiyasasi@mepasenerji.com']
port = 25
msg=MIMEMultipart()
msg['Subject'] = 'Random Forest K3 Tahmini'
msg['From'] = 'info@mepasenerji.com'
msg['To'] = 'enerjipiyasasi@mepasenerji.com'
eklenti_dosya_ismi=z
dsgFilename='tahmin.xlsx'
msg.attach(MIMEText("merhaba K3 Random Forest tahmini ektedir.Saygılarımla"))
with(open(eklenti_dosya_ismi,'rb')) as eklenti_dosyasi:
    payload=MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet',name=dsgFilename)
    payload.set_payload(eklenti_dosyasi.read())
    encoders.encode_base64(payload)
    payload.add_header("Content-Decomposition","attachment",filename=eklenti_dosya_ismi)
    msg.attach(payload)
    msg_str=msg.as_string()
with smtplib.SMTP('mail.mepasenerji.com', 25) as server:
    server.sendmail(sender, receivers, msg_str)
    print("Successfully sent email")


# In[ ]:


import os
os.remove(z, dir_fd=None)


# In[ ]:




