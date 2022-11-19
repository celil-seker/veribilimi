#!/usr/bin/env python
# coding: utf-8

# # XGBOOST ALGORİTMASI

# In[2]:


import xgboost as xgb


# In[3]:


import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error, r2_score
from sklearn import model_selection
from sklearn.tree import DecisionTreeRegressor
import openpyxl


# In[4]:


from warnings import filterwarnings
filterwarnings("ignore")


# In[ ]:


df= pd.read_excel('//MPSTLPSRV2//TalepTahminPaylasim2//model//verison.xlsx')
df1=pd.read_excel('//MPSTLPSRV2//TalepTahminPaylasim2/model//üretim.xlsx')


# In[ ]:


df.info()


# In[ ]:


df1.info()


# In[ ]:


toplam=0
toplaagg=0
aksaray_agırlık=0.1
konya_agırlık=0.5
nigde_agırlık=0.1
karaman_agırlık=0.1
nevsehir_agırlık=0.1
kırsehir_agırlık=0.1
toplaagırlık=aksaray_agırlık+nigde_agırlık+konya_agırlık+karaman_agırlık+nevsehir_agırlık+kırsehir_agırlık
print(aksaray_agırlık)
print(nigde_agırlık)
print(konya_agırlık)
print(karaman_agırlık)
print(nevsehir_agırlık)
print(kırsehir_agırlık)


# In[ ]:


toplaagırlık


# In[ ]:


k=df[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]


# In[ ]:


sehir=df["Şehir"]
y_üret=df["Radiation"]


# In[ ]:


for v in range(len(y_üret.index)): 
    if (sehir[v] == 'AKSARAY') :
        k[v:v+1]=k[v:v+1]*aksaray_agırlık    
    elif(sehir[v] == 'NİĞDE') :  
        k[v:v+1]=k[v:v+1]*nigde_agırlık
    elif(sehir[v] == 'KONYA') :
        k[v:v+1]=k[v:v+1]*konya_agırlık
    elif(sehir[v] == 'KARAMAN') : 
        k[v:v+1]=k[v:v+1]*karaman_agırlık
    elif(sehir[v] == 'KIRŞEHİR') :
        k[v:v+1]=k[v:v+1]*kırsehir_agırlık
    elif(sehir[v] == 'NEVŞEHİR') :
        k[v:v+1]=k[v:v+1]*nevsehir_agırlık


# In[ ]:


df[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]=k[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]
table = pd.pivot_table(df,index=['Tarih'],aggfunc=np.sum)


# In[ ]:


df1.index=table.index


# In[ ]:


table=table.join(df1)


# In[ ]:


y = table.loc[:,["GES NET"]]
c=table.loc[:,["Radiation","Effective Cloud","Medium Cloud","Solar Power","Precipitation","Low Cloud","Clear Radiation","Temperature"]]
print('X Shape: ', c.shape)
print('Y Shape: ', y.shape)


# In[ ]:


c_train, c_test, y_train, y_test = train_test_split(c,
                                                    y,
                                                    test_size = 0.00015,
                                                    random_state= 42)


# In[ ]:


xg_reg = xgb.XGBRegressor(objective ='reg:squarederror', colsample_bytree=0.9, eta=0.9, 
                              gamma=15, max_depth=200, n_estimators=10000, subsample=0.6,learning_rate=0.6,random_state=42, reg_alpha=5, reg_lambda=10,min_child_weight=3)


# In[ ]:


model=xg_reg.fit(c_train,y_train)


# In[1]:


import pypyodbc
db = pypyodbc.connect(
   'Driver={SQL Server};'
    'Server=10.242.1.135;'
    'Database=Mepas;'
    'UID=TalepTahminGoruntuleme;'
    'PWD=Mepas@951;')
imlec = db.cursor()


# In[16]:


df = pd.read_sql_query("SELECT Sehir,Tarih,tip, Sirket,Temperature,Precipitation,WindSpeed,LowCloud,MediumCloud,HighCloud,EffectiveCloud,Radiation,ClearRadiation,SolarPower From K3HavaTahmini",db)


# In[17]:


hava_tahmin=df[df['tip'] == 'Tahmin']


# In[18]:


hava_tahmin.rename(columns ={'sehir':'Şehir','tarih':'Tarih','radiation':'Radiation','temperature':'Temperature','precipitation':'Precipitation','windspeed':'Wind Speed','lowcloud':'Low Cloud','mediumcloud':'Medium Cloud','highcloud':'High Cloud','effectivecloud':'Effective Cloud','clearradiation':'Clear Radiation','solarpower':'Solar Power'},inplace=True)


# In[ ]:


hava_tahmin.reset_index(drop=True,inplace=True)


# In[ ]:


sehir_tah=hava_tahmin["Şehir"]
ke=hava_tahmin[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]
for v in range(len(hava_tahmin.index)): 
    if (sehir_tah[v] == 'AKSARAY') :
        ke[v:v+1]=ke[v:v+1]*aksaray_agırlık
    elif(sehir_tah[v] == 'NİĞDE') :
        ke[v:v+1]=ke[v:v+1]*nigde_agırlık
    elif(sehir_tah[v]  == 'KONYA') :
        ke[v:v+1]=ke[v:v+1]*konya_agırlık
    elif(sehir_tah[v]  == 'KARAMAN') :
        ke[v:v+1]=ke[v:v+1]*karaman_agırlık
    elif(sehir_tah[v]  == 'KIRŞEHİR') :
        ke[v:v+1]=ke[v:v+1]*kırsehir_agırlık
    elif(sehir_tah[v]  == 'NEVŞEHİR') :
        ke[v:v+1]=ke[v:v+1]*nevsehir_agırlık


# In[ ]:


hava_tahmin[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]=ke[["Radiation","Solar Power","Temperature","Precipitation","Wind Speed","Low Cloud", "Medium Cloud","High Cloud","Effective Cloud","Clear Radiation"]]


# In[ ]:


table_tah = pd.pivot_table(hava_tahmin,index=['Tarih'],aggfunc=np.sum)


# In[ ]:


c_tahmin=table_tah.loc[:,["Radiation","Effective Cloud","Medium Cloud","Solar Power","Precipitation","Low Cloud","Clear Radiation","Temperature","High Cloud","Wind Speed"]]
c_tahmin=c_tahmin.reset_index(drop=True)


# In[22]:


import datetime
tarih =datetime.datetime.now()
zaman_damgası = datetime.datetime.timestamp(tarih)
zaman_damgası
ce=str(zaman_damgası) + ".xlsx"


# In[ ]:


wb = openpyxl.Workbook()
sayfa = wb.active
for v in range(len(c_tahmin["Radiation"].index)):     
    y_pred =model.predict(c_tahmin.loc[[v],["Radiation","Effective Cloud","Medium Cloud","Solar Power","Precipitation","Low Cloud","Clear Radiation","Temperature"]]) 
    r = v + 1
    sayfa.cell(row = r, column = 2).value = float(y_pred)
    sayfa.cell(row = r, column = 1).value = table_tah.index[v]
wb.save(ce)
wb.close()


# In[ ]:


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
sender = 'info@mepasenerji.com'
receivers = ['enerjipiyasasi@mepasenerji.com']
port = 25
msg=MIMEMultipart()
msg['Subject'] = 'Xgboost K3 Tahmini'
msg['From'] = 'info@mepasenerji.com'
msg['To'] = 'enerjipiyasasi@mepasenerji.com'
eklenti_dosya_ismi=ce
dsgFilename='tahmin.xlsx'
msg.attach(MIMEText("merhaba K3 Xgboost  tahmini ektedir.Saygılarımla"))
with(open(eklenti_dosya_ismi,'rb')) as eklenti_dosyasi:
    payload=MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet',name=dsgFilename)
    payload.set_payload(eklenti_dosyasi.read())
    encoders.encode_base64(payload)
    payload.add_header("Content-Decomposition","attachment",filename=eklenti_dosya_ismi)
    msg.attach(payload)
    msg_str=msg.as_string()
with smtplib.SMTP('mail.mepasenerji.com', 25) as server:
    server.sendmail(sender, receivers, msg_str)
    print("Successfully sent email")


# In[ ]:


import os
os.remove(ce, dir_fd=None)


# In[3]:


import os
import platform
from typing import List
import sys


class XGBoostLibraryNotFound(Exception):
    """Error thrown by when xgboost is not found"""


def find_lib_path() -> List[str]:
    """Find the path to xgboost dynamic library files.
    Returns
    -------
    lib_path
       List of all found library path to xgboost
    """
    curr_path = os.path.dirname(os.path.abspath(os.path.expanduser(__file__)))
    dll_path = [
        # normal, after installation `lib` is copied into Python package tree.
        os.path.join(curr_path, 'lib'),
        # editable installation, no copying is performed.
        os.path.join(curr_path, os.path.pardir, os.path.pardir, 'lib'),
        # use libxgboost from a system prefix, if available.  This should be the last
        # option.
        os.path.join(sys.prefix, 'lib'),
    ]

    if sys.platform == 'win32':
        if platform.architecture()[0] == '64bit':
            dll_path.append(
                os.path.join(curr_path, '../../windows/x64/Release/'))
            # hack for pip installation when copy all parent source
            # directory here
            dll_path.append(os.path.join(curr_path, './windows/x64/Release/'))
        else:
            dll_path.append(os.path.join(curr_path, '../../windows/Release/'))
            # hack for pip installation when copy all parent source
            # directory here
            dll_path.append(os.path.join(curr_path, './windows/Release/'))
        dll_path = [os.path.join(p, 'xgboost.dll') for p in dll_path]
    elif sys.platform.startswith(('linux', 'freebsd', 'emscripten')):
        dll_path = [os.path.join(p, 'libxgboost.so') for p in dll_path]
    elif sys.platform == 'darwin':
        dll_path = [os.path.join(p, 'libxgboost.dylib') for p in dll_path]
    elif sys.platform == 'cygwin':
        dll_path = [os.path.join(p, 'cygxgboost.dll') for p in dll_path]
    if platform.system() == 'OS400':
        dll_path = [os.path.join(p, 'libxgboost.so') for p in dll_path]

    lib_path = [p for p in dll_path if os.path.exists(p) and os.path.isfile(p)]

    # XGBOOST_BUILD_DOC is defined by sphinx conf.
    if not lib_path and not os.environ.get('XGBOOST_BUILD_DOC', False):
        link = 'https://xgboost.readthedocs.io/en/latest/build.html'
        msg = 'Cannot find XGBoost Library in the candidate path.  ' +             'List of candidates:\n- ' + ('\n- '.join(dll_path)) +             '\nXGBoost Python package path: ' + curr_path +             '\nsys.prefix: ' + sys.prefix +             '\nSee: ' + link + ' for installing XGBoost.'
        raise XGBoostLibraryNotFound(msg)
    return lib_path


# In[ ]:




